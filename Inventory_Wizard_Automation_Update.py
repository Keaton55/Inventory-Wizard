import csv
from openpyxl import Workbook
import openpyxl
import datetime
import re
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from pydomo import Domo
import win32com.client
from pathlib import Path
import shutil
import os
import sys
import numpy as np
import pandas as pd
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

def main():

    Recipients = 'Keaton Manwaring <keaton@blackstoneproducts.com>'
    #Recipients = 'Kaden Merrill <kaden.merrill@blackstoneproducts.com>;Christian Elkins<christian@blackstoneproducts.com>;Steven Bassett<steveb@blackstoneproducts.com>;Chloe Bowman<chloe@blackstoneproducts.com>; Canon Schenk<Canon.schenk@blackstoneproducts.com>; Cameron Gardner<cameron.gardner@blackstoneproducts.com>;Kalin Hansen<kalin.hansen@blackstoneproducts.com>;Abby Griffeth<abbyg@blackstoneproducts.com>;Chris  Tucket <tchris@blackstoneproducts.com>;Spencer Stratton <spencer@blackstoneproducts.com>;Aaron Smart <aaron@blackstoneproducts.com>; Logan Rondash <logan@blackstoneproducts.com>; Chris Brown <chris@blackstoneproducts.com>; Darren Cole <darren@blackstoneproducts.com>; David Anderson <davida@blackstoneproducts.com>; Joyce Jensen <Joyce@blackstoneproducts.com>; Keaton Manwaring <keaton@blackstoneproducts.com>; Kjersti Green <kgreen@blackstoneproducts.com>; Kyler Hansen <kyler@blackstoneproducts.com>; Mark Malen <mark@blackstoneproducts.com>; Michael Jenkins <mjenkins@blackstoneproducts.com>; Mike Midgley <mike@blackstoneproducts.com>; Mike Moser <mike.moser@blackstoneproducts.com>; Perry Jensen <Perry@blackstoneproducts.com>; Ty H <ty@blackstoneproducts.com>; Tom Newman <tom@blackstoneproducts.com>; Nicholle Anderson <nicholle@blackstoneproducts.com>; Clayton Shaw <clayton@blackstoneproducts.com>;  Holley Creger <Holley@blackstoneproducts.com>; Venessa P <venessa@blackstoneproducts.com>; Travis Cox <travis@blackstoneproducts.com>; Jake D <Jake@blackstoneproducts.com>; Brad Wheelwright <brad@blackstoneproducts.com>; Jared Jensen <jj@blackstoneproducts.com>; Tann Tueller <tann@blackstoneproducts.com>; Trevor  Gonzalez <trevor@blackstoneproducts.com>; Vance  Jensen <vance@blackstoneproducts.com>; Import <import@blackstoneproducts.com>; Ian Dahle <ian@blackstoneproducts.com> ; Keaton Manwaring <keaton@blackstoneproducts.com>'
    email_pattern = r'<([^>]+)>'
    email_list = re.findall(email_pattern, Recipients)

    domo = Domo('21cb74ca-7f01-4fc2-98ff-4458de43561b','44434b52cad91ad48371774134911eff6c6d2a5f8fa582f764ee5a5c31ddb5ef')

    #Inventory Wizard Dataflow
    Wizard = domo.ds_query('57c287af-ca21-42e0-8c02-21bc7d025356','select * from table')
    #Wizard = domo.ds_get('57c287af-ca21-42e0-8c02-21bc7d025356')

    Wizard = Wizard[Wizard['Type'] != 'Fulfillments']
    Wizard = Wizard[Wizard['Type'] != 'Inventory Locations']
    Wizard['Item'] = Wizard['Item'].astype(str)
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('RP', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('misc', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('Dollar', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('Grease', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('Sales Discount', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('Colorado', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('90424', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('90252', na=False)]
    Wizard = Wizard[ ~ Wizard['Item'].str.contains('US#ams37', na=False)]
    Wizard.reset_index(inplace = True)
    Wizard['Schedule Date'] = pd.to_datetime(Wizard['Schedule Date'])


    Inventory = domo.ds_get('13de9dcb-ed35-4a30-a7ee-8e1e88ddceb6')

    #Inventory_CG
    Netsuite_Inventory = domo.ds_get('81246892-3262-4180-8acc-e4c6bac340db')

    #Salsify Data Dataflow
    Item_Information = domo.ds_get('9bac3cfd-d94e-4e25-9baf-d546cb8236c6')

    Shipments = domo.ds_get('244814e5-bd45-45d0-bb77-5f13eb18ce59')


    #Wizard = []
    #Inventory = []
    #Netsuite_Inventory = []
    #Item_Information = []
    #Shipments = []
    #Lunar = []

    # Read Inventory Wizard File From Domo
    """
    with open('Inventory_Wizard.csv','r',encoding='utf8') as wizard:
        reader = csv.DictReader(wizard)
        for row in reader:
            Wizard.append({
                'Customer':row['Customer'],'P.O. Number':row['P.O. Number'],'NAI P.O.':row['NAI P.O.'],'Container #':row['Container #'],
                'Document Number': row['Document Number'], 'Sell Price': row['Sell Price'], 'Item':row['Item'],
                'Location':row['Location'], 'Schedule Date' : row['Schedule Date'],'Quantity':row['Quantity'],
                'Actual Inventory Level': row['Actual Inventory Level'], 'Type': row['Type'], 'Status': row['Status'],
                'Initial Forecast':row['Initial Forecast'],'Forecast Remaining':row['Forecast Remaining'],'Forecast':row['Forecast'],'Forecast Inventory Level': row['Forecast Inventory Level'],
                'Sort':row['Sort'],
                })
    """
        
    """
    #Read Inventory in appendix
    with open('Deposco Inventory.csv','r') as inventory:
        reader = csv.DictReader(inventory)
        for row in reader: 
            Inventory.append({
                'Item': row['Item'],'Split/Full':row['Split/Full'],'Allocated': row['Allocated'] ,'Pickable' : row['Pickable'], 'Storage' : row['Storage'], 'Staged' : row['Staged'],
                'Damaged' : row['Damaged'], 'Rework' : row['Rework'],  'Total' : row['Total'], 'Final' : row['Final'],
            })
    """
    """
    #Read Netsuite Inventory in appendix
    with open('Inventory.csv','r') as inventory:
        reader = csv.DictReader(inventory)
        for row in reader:
            Netsuite_Inventory.append({
                'Item' : row['Item'], 'Location' : row['Location'], 'On Hand' : row['On Hand'],
            })
    """
    """
    #Read Item_information
    with open('Item_Information.csv','r') as item_information:
        reader = csv.DictReader(item_information)
        for row in reader: 
            Item_Information.append({
                'Unique ID': row['Unique ID'], 'Display Name': row['Display Name'],'Product Group': row['Product Group'], 'Program Year': row['Program Year'], 
                'Item Status': row['Item Status'], 'Master Pack Quantity' : row['Master Pack Quantity'], 'Exclusivity': row['Exclusivity'],'Factory':row['Factory'],
                'Old SKU' : row['Old SKU'], 'New SKU' : row['New SKU'],'Last Purchase Price':row['Last Purchase Price'] , 'Purchase Price': row['Purchase Price'],
                'Unclaimed Inventory' : row['Unclaimed Inventory'],'Date Last Updated' : row['Date Last Updated'], "QTY PER 40'HQ CONTAINER" : row["QTY PER 40'HQ CONTAINER"],
            })
    """
    """
    #Read Shipment information
    with open('Shipments.csv','r') as shipments:
        reader = csv.DictReader(shipments)
        for row in reader: 
            Shipments.append({
                'Item':row['Item'], 'Import/Domestic': row['Import/Domestic'],'LY Quantity': row['LY Quantity'], 'YTD Quantity': row['YTD Quantity'], 
                'MTD Quantity': row['MTD Quantity'], 'LY Cancelled' : row['LY Cancelled'], 'YTD Cancelled' : row['YTD Cancelled'],'MTD Cancelled' : row['MTD Cancelled']
                , 'LY YTD Quantity' : row['LY YTD Quantity']
            })
    """
    #Read Factory Lunar New Year Holiday From Excel
    """
    with open('Factory Lunar New Year Holday Information.csv') as lunar:
        reader = csv.DictReader(lunar)
        for row in reader:
            Lunar.append({
                'Factory Name' : row['Factory Name'], 'Prodcution Country' : row['Production Country'], 'Loading Port' : row['Loading Port'], 
                'Ship Date before CNY' : row['Best Ship Date for the Last Shipment  before CNY'], 'Order Date before CNY' : row['Expected order place date for Last pre-CNY order'],
                'Factory Open Date' : row['Factory Opening Date after CN'], 'Ship Date after CNY' : row['Best Ship Date for the 1st Shipment after CNY'],
                'Order Date after CNY' : row['Expected order place date for the 1st post CNY order']
            })
    """

    #create file name
    book_name = 'NAI_Inventory_Wizard' + ' ' + str(datetime.datetime.today().strftime("%Y-%m-%d %A %H;%M;%S"))+".xlsx"

    #date = datetime.date.isoformat(datetime.date.today())
    #time = datetime.datetime.today().__format__('%H:%M:%S')

    #create workbook and unique SKU list as well as sort
    wb = Workbook()
    unique_list = []

    for item in range(len(Wizard.sort_values('Item'))):
        try:
            if Wizard['Customer'][item] == 'Beginning Inventory' and int(float(Wizard['Quantity Remaining'][item])) == 0:
                pass
            elif Wizard['Item'][item] not in unique_list:
                unique_list.append(Wizard['Item'][item])
        except ValueError:
            pass
    


    #create common formatting techniques for use throughout
    
    thin_border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
    #thick_border = openpyxl.styles.borders.Border(
                #left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thick'),
                #top=openpyxl.styles.borders.Side(style='thick')
                #)
    center = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold = openpyxl.styles.Font(bold=True)
    underline = openpyxl.styles.Font(underline = 'single')
    double_underline = openpyxl.styles.Font(underline = 'double')
    right = openpyxl.styles.Alignment(horizontal='right', vertical='center', wrap_text=False)
    italic = openpyxl.styles.Font(italic = True)

    

    #Read header names
    all_keys = [
    'Customer','P.O. Number', 'NAI P.O','Container #', 'Document Number','Sell Price', 'Item','Location','Schedule Date', 'Quantity','Actual Inventory Level',
    'Type', 'Status','Initial Forecast', 'Forecast Remaining','Forecast','Forecast Inventory Level'
    ]

    Wizard['P.O. Number'] = np.where(pd.isnull(Wizard['P.O. Number']),Wizard['P.O. Number'],Wizard['P.O. Number'].astype(str))
    #Wizard.replace(np.nan,"")
    #Wizard['P.O. Number'] = Wizard['P.O. Number'].astype('str')

    #start of worksheet 
    for SKU in unique_list:
        ws = wb.create_sheet(title=SKU)
        count = 1 
        #header rows created here 
        header_row = 13
        for i in all_keys:
            ws.cell(row=header_row, column = count, value = i)
            ws.cell(row=header_row, column=count).alignment = center
            ws.cell(row=header_row, column=count).font = bold
            ws.cell(row=header_row, column=count).border = thin_border
            count += 1
        
        ws.cell(row=header_row, column = count,value = 'Inventory Adjustments').alignment = center
        ws.cell(row = header_row, column = count).font = bold
        ws.cell(row= header_row, column = count).border = thin_border

        #column widths set here
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 11
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 25
        ws.column_dimensions['M'].width = 17
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 11
        ws.column_dimensions['P'].width = 11
        ws.column_dimensions['Q'].width = 11
        ws.column_dimensions['R'].width = 28
        ws.column_dimensions['S'].width = 15
        #ws.column_dimensions['T'].width = 15

        ws.freeze_panes = 'A14'
        ws.sheet_view.zoomScale = 90
        filter = ws.auto_filter
        filter.ref = f"A{header_row}:S{header_row}"

        current_row = header_row + 1
        #Sorting

        #def catch(func,*args,**kwargs):
        #    try:
        #        return func(*args,**kwargs)
        #    except:
        #        return 0 
        
        #Wizard['P.O. Number Sort'] = [catch(lambda : int(i.split(r'[a-zA-Z]')[-1])) for i in Wizard['P.O. Number']]
        #Wizard['NAI P.O. Sort'] = [catch(lambda : int(i.split(r'[a-zA-Z]')[-1])) for i in Wizard['NAI P.O.']]

        Wizard['NAI P.O. Sort'] = pd.to_numeric(Wizard['NAI P.O.'].str.replace('PO',''),errors='coerce')


        sort = Wizard.sort_values('NAI P.O. Sort')
        sort = sort.sort_values('Schedule Date')
        sort = sort.sort_values('Beginning Inventory?',ascending = False)
        #Start of Item Specifics 
        for row in range(len(sort)): 
            if sort['Item'][row] == SKU:
                #convert Text to Numbers
                try:
                    sell_price = float(sort['Sell Price'][row])
                except ValueError:
                    sell_price = sort['Sell Price'][row]
                try:
                    Item = int(float(sort['Item'][row]))
                except ValueError:
                    Item = sort['Item'][row]
                try:
                    Quantity = int(float(sort['Quantity Remaining'][row]))
                except ValueError:
                    Quantity = sort['Quantity Remaining'][row]
                try:
                    Actual_Inventory_Level = int(float(sort['Actual Inventory Level'][row]))
                except ValueError:
                    Actual_Inventory_Level = sort['Actual Inventory Level'][row]

                try:
                    Initial_Forecast = int(float(sort['Original Forecast Qty'][row]))
                except ValueError:
                    Initial_Forecast = sort['Original Forecast Qty'][row]
                try:
                    Forecast_Remaining = int(float(sort['Monthly Quantities'][row])) * -1
                except ValueError:
                    Forecast_Remaining = sort['Monthly Quantities'][row]
                try:
                    Forecast = int(float(sort['Forecast'][row]))
                except ValueError:
                    Forecast = sort['Forecast'][row]
                try:
                    if sort['Type Column'][row] == 'Beginning Inventory': 
                        Forecast_Inventory_Level = int(float(sort['Forecast Inventory Level'][row]))
                    else: 
                        
                        Forecast_Inventory_Level = '=IF(L'+str(current_row)+'="P.O.",Q'+str(current_row-1)+',IF(H'+str(current_row)+'="Import",Q'+str(current_row-1)+',IF(L'+str(current_row)+'="Forecast",Q'+str(current_row-1)+'+(P'+str(current_row)+'*$Q$12),Q'+str(current_row-1)+'+P'+str(current_row)+'))) + R' + str(current_row)
                
                except ValueError:
                    Forecast_Inventory_Level = sort['Forecast Inventory Level'][row]   

                ws.cell(row=header_row,column = 19,value='Memo')
                ws.cell(row=header_row, column=19).alignment = center
                ws.cell(row=header_row, column=19).font = bold
                ws.cell(row=header_row, column=19).border = thin_border  
                try: 
                    ws.cell(row=current_row, column=19, value=sort['Memo (Main)'][row])
                except ValueError: 
                    pass


                #Asign sort Numbers to cells
                ws.cell(row=current_row, column=1, value=sort['Customer'][row])
                ws.cell(row=current_row, column=2, value=sort['P.O. Number'][row])
                ws.cell(row=current_row, column=3, value=sort['NAI P.O.'][row])
                ws.cell(row=current_row, column=4, value=sort['Container #'][row])
                ws.cell(row=current_row, column=5, value=sort['Document Number'][row])
                ws.cell(row=current_row, column=6, value= sell_price)
                ws.cell(row=current_row, column=7, value= Item)
                ws.cell(row=current_row, column=8, value=sort['Location'][row])
                ws.cell(row=current_row, column=9, value=sort['Schedule Date'][row].strftime('%Y-%m-%d'))
                ws.cell(row=current_row, column=10, value=Quantity)
                ws.cell(row=current_row, column=11, value=Actual_Inventory_Level)
                ws.cell(row=current_row, column=12, value=sort['Type Column'][row])
                ws.cell(row=current_row, column=13, value=sort['Status'][row])
                ws.cell(row=current_row, column=14, value=Initial_Forecast)
                ws.cell(row=current_row, column=15, value=Forecast_Remaining)
                ws.cell(row=current_row, column=16, value=Forecast)
                ws.cell(row=current_row, column=17, value=Forecast_Inventory_Level)
                

                


            # Gray out Forecast Inventory Level Based off of CNY 

                #forecast_belowedit = re.split('-',sort['Schedule Date'][row])
                #forecast_days = datetime.datetime(sort['Schedule Date'][row].year,sort['Schedule Date'][row].month,sort['Schedule Date'][row].day)
                #datetime.datetime.strptime(sort['Schedule Date'][row],'yyyy-mm-dd')
                #CNY_start = datetime.datetime(2024,2,28)
                #CNY_end = datetime.datetime(2024,4,30)
                #if CNY_end > forecast_days > CNY_start:
                    #ws.cell(row=current_row,column = 17).fill = openpyxl.styles.PatternFill(start_color = 'D3D3D3',end_color = 'D3D3D3', fill_type = 'solid')

            # Change the tab color if there are issues 6 months out
                if sort['Schedule Date'][row] <= (datetime.datetime.today() + datetime.timedelta(days = 4 * 30)):
                    if int(float(sort['Forecast Inventory Level'][row])) < 0:
                        ws.sheet_properties.tabColor = 'FFFF00'    


                if sort['Schedule Date'][row].month == (datetime.datetime.today() + datetime.timedelta(days = 4 * 30)).month and sort['Schedule Date'][row].year == (datetime.datetime.today() + datetime.timedelta(days = 4 * 30)).year:
                    if int(float(sort['Forecast Inventory Level'][row])) < 0:
                        ws.sheet_properties.tabColor = '880808'
                
            
        
                current_row += 1

            else:
                pass

        #Sum Inventory positions
        BPUT1 = 0
        BPUT2 = 0
        BPUT3 = 0
        BPUT4 = 0 
        #All_Roads_Trucking = 0
        #Keyword_SLC = 0 
        Clearfield = 0 
        Partners_Trade = 0
        Retail_Store = 0
        HC_Group = 0 
        WFS = 0 

        current_inventory = Netsuite_Inventory[Netsuite_Inventory['Item'] == SKU]
        current_inventory.reset_index(inplace = True)

        for row in range(len(current_inventory)):
            if current_inventory['Location'][row] == 'BPUT1':
                BPUT1 += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'BPUT2':
                BPUT2 += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'BPUT3':
                BPUT3 += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'BPUT4':
                BPUT4 += int(float(current_inventory['On Hand'][row]))
            #elif current_inventory['Location'][row] == 'All Roads Trucking':
            #    All_Roads_Trucking += int(float(current_inventory['On Hand'][row]))
            #elif current_inventory['Location'][row] == 'Keyword SLC':
            #    Keyword_SLC += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'Clearfield':
                Clearfield += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'Partners Trade':
                Partners_Trade += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'Retail Store':
                Retail_Store += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'HC Group':
                HC_Group += int(float(current_inventory['On Hand'][row]))
            elif current_inventory['Location'][row] == 'WFS':
                WFS += int(float(current_inventory['On Hand'][row]))
        


        split_final = 0
        full_final = 0 
        damaged = 0 
        rework = 0 

        current_inventory = Inventory[Inventory['Item'] == SKU]
        current_inventory.reset_index(inplace = True)
        for i in range(len(current_inventory)):
            if current_inventory['Split/Full'][i] == 'Full': 
                try:
                    full_final += int(float(current_inventory['Final'][i]))
                    damaged += int(float(current_inventory['Damaged'][i]))
                    rework += int(float(current_inventory['Rework'][i]))

                except ValueError:
                    pass

            elif current_inventory['Split/Full'][i] == 'Split': 
                try:
                    split_final += int(float(current_inventory['Final'][i]))
                    damaged += int(float(current_inventory['Damaged'][i]))
                    rework += int(float(current_inventory['Rework'][i]))
                except ValueError:
                    pass

        
        #Assign Inventory Values

        ws.cell(row=4,column=8,value='Current Instock by Location:')
        ws.cell(row=4,column=8).font = bold
        ws.cell(row=4,column=8).alignment = right 
        
        ws.cell(row=5,column=6,value='BPUT1:')
        ws.cell(row=5,column=6).font = underline
        ws.cell(row=5,column=6).alignment = right
        ws.cell(row=5,column=7).alignment = center
        ws.cell(row=5,column=7,value=BPUT1)

        ws.cell(row=6,column=6,value='BPUT2:')
        ws.cell(row=6,column=6).font = underline
        ws.cell(row=6,column=6).alignment = right
        ws.cell(row=6,column=7).alignment = center
        ws.cell(row=6,column=7,value=BPUT2)

        ws.cell(row=7,column=6,value='BPUT3:')
        ws.cell(row=7,column=6).font = underline
        ws.cell(row=7,column=6).alignment = right
        ws.cell(row=7,column=7).alignment = center
        ws.cell(row=7,column=7,value=BPUT3)

        ws.cell(row=5,column=9,value='BPUT4:')
        ws.cell(row=5,column=9).font = underline
        ws.cell(row=5,column=9).alignment = right
        ws.cell(row=5,column=10).alignment = center
        ws.cell(row=5,column=10,value=BPUT4)

        ws.cell(row=6,column=9,value='Retail Store:')
        ws.cell(row=6,column=9).font = underline
        ws.cell(row=6,column=9).alignment = right
        ws.cell(row=6,column=10).alignment = center
        ws.cell(row=6,column=10,value=Retail_Store)

        ws.cell(row=9,column=9,value='WFS:')
        ws.cell(row=9,column=9).font = underline
        ws.cell(row=9,column=9).alignment = right
        ws.cell(row=9,column=10).alignment = center
        ws.cell(row=9,column=10,value=WFS)

        ws.cell(row=7,column=9,value='Clearfield:')
        ws.cell(row=7,column=9).font = underline
        ws.cell(row=7,column=9).alignment = right
        ws.cell(row=7,column=10).alignment = center
        ws.cell(row=7,column=10,value=Clearfield)

        ws.cell(row=8,column=9,value='HC Group:')
        ws.cell(row=8,column=9).font = underline
        ws.cell(row=8,column=9).alignment = right
        ws.cell(row=8,column=10).alignment = center
        ws.cell(row=8,column=10,value=HC_Group)

        ws.cell(row=5,column=12,value='Partners Trade:')
        ws.cell(row=5,column=12).font = underline
        ws.cell(row=5,column=12).alignment = right
        ws.cell(row=5,column=13).alignment = center
        ws.cell(row=5,column=13,value=Partners_Trade)

        ws.cell(row=6,column=12,value='Damaged Location:')
        ws.cell(row=6,column=12).font = underline
        ws.cell(row=6,column=12).alignment = right
        ws.cell(row=6,column=13).alignment = center
        ws.cell(row=6,column=13,value=damaged)

        ws.cell(row=7,column=12,value='Rework:')
        ws.cell(row=7,column=12).font = underline
        ws.cell(row=7,column=12).alignment = right
        ws.cell(row=7,column=13).alignment = center
        ws.cell(row=7,column=13,value=rework)

        ws.cell(row=10,column=6,value='Split:')
        ws.cell(row=10,column=6).font = underline
        ws.cell(row=10,column=6).alignment = right
        ws.cell(row=10,column=7).alignment = center
        ws.cell(row=10,column=7,value=split_final)

        ws.cell(row=9,column=6,value='Full:')
        ws.cell(row=9,column=6).font = underline
        ws.cell(row=9,column=6).alignment = right
        ws.cell(row=9,column=7).alignment = center
        ws.cell(row=9,column=7,value=full_final)
        
        #Asign Item Information to cells
        ws.cell(row=1,column=1,value='Item Number:') 
        ws.cell(row=1,column=1).alignment = center

        ws.cell(row=1,column=2,value=SKU)
        ws.cell(row=1,column=2).alignment = center
        ws.cell(row=1,column=2).font = bold

        #Item Information Labels       
        ws.cell(row=4,column=1,value='Product Group:')
        ws.cell(row=5,column=1,value='Program Year:')
        ws.cell(row=6,column=1,value='Item Status:')
        ws.cell(row=7,column=1,value='Master Pack Quantity:')
        ws.cell(row=8,column=1,value='Exclusivity:')
        ws.cell(row=9,column=1,value='Factory:')
        ws.cell(row=11,column=1,value='Import/Domestic:')
        ws.cell(row=12,column=1,value='Packaging Style:')



        Purchase_Price = ""
        Last_Purchase_Price = ""
        unclaimed_inventory = 0 
        Container_Qty = ''
        date_last_updated = ''
        factory = ''

        #Item Information Values
        info = Item_Information[Item_Information['Name'] == SKU]
        info.reset_index(inplace = True)
        for item in range(len(info)):
            try:
                MP_Quantity = str(int(float(info['Master Pack Quantity'][item])))
            except:
                MP_Quantity = info['Master Pack Quantity'][item]
            try:
                Purchase_Price = float(info['Purchase Price'][item])
            except ValueError:
                pass
            try:
                Last_Purchase_Price = float(info['Last Purchase Price'][item])
            except ValueError:
                pass
            try:
                unclaimed_inventory = float(info['Unclaimed Inventory'][item])
            except ValueError:
                pass

            try: 
                date_last_updated = info['Date Last Updated'][item].strftime('%Y-%m-%d')
            except ValueError:
                pass
            
            try: 
                factory = info['Factory'][item]
            except ValueError:
                pass
            

            ws.cell(row=11,column=2,value=info['Import Only'][item])



            ws.cell(row=2,column=1,value=info['Display Name'][item])
            ws.cell(row=4,column=2,value=info['Product Group'][item])
            ws.cell(row=5,column=2,value=info['Program Year'][item])
            ws.cell(row=6,column=2,value=info['Item Status'][item])
            ws.cell(row=7,column=2,value= MP_Quantity)
            ws.cell(row=8,column=2,value=info['Exclusivity'][item])
            ws.cell(row=12,column=2,value=info['Packaging Style'][item])
            ws.cell(row=9,column=2,value=factory)
            ws.cell(row=2,column=1).font = bold
            ws.cell(row=4,column=2).font = bold
            ws.cell(row=5,column=2).font = bold
            ws.cell(row=6,column=2).font = bold
            ws.cell(row=7,column=2).font = bold
            ws.cell(row=8,column=2).font = bold
            ws.cell(row=9,column=2).font = bold
            ws.cell(row=10,column=2).font = bold
            ws.cell(row=11,column=2).font = bold
            ws.cell(row=12,column=2).font = bold

            #Wizard Notes
            ws.cell(row=1,column=11,value='Notes:').font = underline

            ws.cell(row=1,column=12,value=info['Wizard Notes'][item])

            #Unclaimed Inventory
            ws.cell(row=9,column=12,value='Factory Inventory:').font = underline
            ws.cell(row=9,column=12).alignment = right
            ws.cell(row=9,column=13,value = unclaimed_inventory).alignment = center

            ws.cell(row=10,column=12,value='Date Last Updated:').font = underline
            ws.cell(row=10,column=12).alignment = right
            ws.cell(row=10,column=13,value = date_last_updated)

            #40GP Conatiner Qty
            ws.cell(row=10,column=1,value="40 HQ Container Qty")
            ws.cell(row=10,column=2,value=info["QTY PER 40'HQ CONTAINER"][item]).font = bold


            #Old SKU, New SKU
            ws.cell(row=1,column=4,value='Old SKU(s):')
            ws.cell(row=2,column=4,value='New SKU(s):')

            old_skus = re.split(',',str(info['Old SKU'][item]))


            column = 5

            for sku in old_skus:
                ws.cell(row=1,column=column,value=sku).hyperlink = (book_name + '#' + sku + '!' + 'A1')
                column += 1

            new_skus = re.split(',',str(info['New SKU'][item]))


            column = 5
            
            for sku in new_skus:
                ws.cell(row=2,column=column,value=sku).hyperlink = (book_name + '#' + sku + '!' + 'A1')
                column += 1


        #Asign Purchase Price and Last Purchase Price to cells 
        ws.cell(row=10,column=4,value='Purchase Price').alignment = right
        ws.cell(row=11,column=4,value='Last Purchase Price').alignment = right

        ws.cell(row=10,column=5,value=Purchase_Price)
        ws.cell(row=11,column=5,value=Last_Purchase_Price)

        #Shipment Labels
        ws.cell(row=3,column=15,value='Import')
        ws.cell(row=3,column=15).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
        ws.cell(row=3,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=3,column=16,value='Domestic')
        ws.cell(row=3,column=16).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
        ws.cell(row=3,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=3,column=17,value='Total')
        ws.cell(row=3,column=17).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
        ws.cell(row=3,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=4,column=14,value='LY Shipped')
        ws.cell(row=4,column=14).alignment = center
        #ws.cell(row=9,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=4,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=14,value='LY YTD Shipped')
        ws.cell(row=5,column=14).alignment = center
        #ws.cell(row=10,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=5,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=6,column=14,value='YTD Shipped')
        ws.cell(row=6,column=14).alignment = center
        #ws.cell(row=10,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=6,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=7,column=14,value='MTD Shipped')
        ws.cell(row=7,column=14).alignment = center
        #ws.cell(row=11,column=13).font = openpyxl.styles.Font(size=8)
        ws.cell(row=7,column=14).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=8,column=14,value='LY Cancelled').border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=8,column=14).alignment = center
        ws.cell(row=9,column=14,value='YTD Cancelled').border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=9,column=14).alignment = center
        ws.cell(row=10,column=14,value='MTD Cancelled').border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thick'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=10,column=14).alignment = center


        #Assign Shipment Values
        import_LY = 0
        import_YTD = 0 
        import_MTD = 0 
        domestic_LY = 0 
        domestic_YTD = 0 
        domestic_MTD = 0 
        total_LY = 0 
        total_YTD = 0 
        total_MTD = 0 
        import_LY_C = 0
        import_YTD_C = 0 
        import_MTD_C = 0 
        domestic_LY_C = 0
        domestic_YTD_C = 0
        domestic_MTD_C = 0 
        total_LY_C = 0 
        total_YTD_C = 0 
        total_MTD_C = 0 
        import_ly_ytd = 0
        domestic_ly_ytd = 0 
        total_ly_ytd = 0 


        current_shipments = Shipments[Shipments['Item'] == SKU]
        current_shipments.reset_index(inplace = True)
        for shipment in range(len(current_shipments)): 
            if current_shipments['Import/Domestic'][shipment] == 'Import':
                try: 
                    import_LY += int(float(current_shipments['LY Quantity'][shipment]))
                except ValueError:
                    pass
                try: 
                    import_YTD += int(float(current_shipments['YTD Quantity'][shipment]))
                except ValueError:
                    pass
                try: 
                    import_MTD += int(float(current_shipments['MTD Quantity'][shipment]))
                except ValueError:
                    pass
                try:
                    import_LY_C += int(float(current_shipments['LY Cancelled'][shipment]))
                except ValueError:
                    pass
                try:
                    import_YTD_C += int(float(current_shipments['YTD Cancelled'][shipment]))
                except ValueError:
                    pass
                try:
                    import_MTD_C += int(float(current_shipments['MTD Cancelled'][shipment]))
                except ValueError:
                    pass
                try:
                    import_ly_ytd += int(float(current_shipments['LY YTD Quantity'][shipment]))
                except ValueError:
                    pass
            else:
                try: 
                    domestic_LY += int(float(current_shipments['LY Quantity'][shipment]))
                except ValueError:
                    pass
                try:
                    domestic_YTD += int(float(current_shipments['YTD Quantity'][shipment]))
                except ValueError:
                    pass
                try:
                    domestic_MTD += int(float(current_shipments['MTD Quantity'][shipment]))
                except ValueError:
                    pass
                try:
                    domestic_LY_C += int(float(current_shipments['LY Cancelled'][shipment]))
                except: 
                    pass
                try:
                    domestic_YTD_C += int(float(current_shipments['YTD Cancelled'][shipment]))
                except:
                    pass
                try:
                    domestic_MTD_C += int(float(current_shipments['MTD Cancelled'][shipment]))
                except:
                    pass
                try:
                    domestic_ly_ytd += int(float(current_shipments['LY YTD Quantity'][shipment]))
                except:
                    pass
            
            total_LY = import_LY + domestic_LY
            total_YTD = import_YTD + domestic_YTD
            total_MTD = import_MTD + domestic_MTD
            total_LY_C = import_LY_C + domestic_LY_C
            total_YTD_C = import_YTD_C + domestic_YTD_C
            total_MTD_C = import_MTD_C + domestic_MTD_C
            total_ly_ytd = import_ly_ytd + domestic_ly_ytd

        #Asign Shipment Values to Cells
        ws.cell(row=4,column=15,value=import_LY).alignment = center
        ws.cell(row=5,column=15,value=import_ly_ytd).alignment = center
        ws.cell(row=6,column=15,value=import_YTD).alignment = center
        ws.cell(row=7,column=15,value=import_MTD).alignment = center  
        ws.cell(row=4,column=16,value=domestic_LY).alignment = center
        ws.cell(row=5,column=16,value=domestic_ly_ytd).alignment = center  
        ws.cell(row=6,column=16,value=domestic_YTD).alignment = center    
        ws.cell(row=7,column=16,value=domestic_MTD).alignment = center 
        ws.cell(row=4,column=17,value=total_LY).alignment = center
        ws.cell(row=5,column=17,value=total_ly_ytd).alignment = center   
        ws.cell(row=6,column=17,value=total_YTD).alignment = center    
        ws.cell(row=7,column=17,value=total_MTD).alignment = center

        ws.cell(row=8,column=15,value=import_LY_C).alignment = center
        ws.cell(row=8,column=16,value=domestic_LY_C).alignment = center  
        ws.cell(row=8,column=17,value=total_LY_C).alignment = center
        ws.cell(row=8,column=15).border = thin_border
        ws.cell(row=8,column=16).border = thin_border
        ws.cell(row=8,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        ws.cell(row=9,column=15,value=import_YTD_C).alignment = center
        ws.cell(row=9,column=16,value=domestic_YTD_C).alignment = center  
        ws.cell(row=9,column=17,value=total_YTD_C).alignment = center
        ws.cell(row=9,column=15).border = thin_border
        ws.cell(row=9,column=16).border = thin_border
        ws.cell(row=9,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        ws.cell(row=10,column=15,value=import_MTD_C).alignment = center
        ws.cell(row=10,column=16,value=domestic_MTD_C).alignment = center  
        ws.cell(row=10,column=17,value=total_MTD_C).alignment = center
        ws.cell(row=10,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=10,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=10,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        ws.cell(row=4,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=15).border = thin_border
        ws.cell(row=6,column=15).border = thin_border
        ws.cell(row=7,column=15).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=4,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=16).border = thin_border
        ws.cell(row=6,column=16).border = thin_border
        ws.cell(row=7,column=16).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thin'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=4,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thick')
                )
        ws.cell(row=5,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=6,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thin'),
                top=openpyxl.styles.borders.Side(style='thin')
                )
        ws.cell(row=7,column=17).border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='thin'),right=openpyxl.styles.borders.Side(style='thick'),bottom=openpyxl.styles.borders.Side(style='thick'),
                top=openpyxl.styles.borders.Side(style='thin')
                )

        #Asign Summary Numbers 
        onhand = 0
        domestic_SO = 0 
        import_SO = 0 
        domestic_PO = 0
        import_PO = 0
        transit = 0 
        available = 0
        import_forecast = 0 
        domestic_forecast = 0
        location_in_transit = 0 

        current_sku = Wizard[Wizard['Item']==SKU]
        current_sku.reset_index(inplace = True)
        for row in range(len(current_sku)):
            if current_sku['Type Column'][row] == 'Beginning Inventory':                
                try: 
                    onhand = int(float(current_sku['Quantity Remaining'][row]))
                except ValueError:
                    onhand = current_sku['Quantity Remaining'][row]

            elif current_sku['Type Column'][row].__contains__('S.O.'):
                if current_sku['Import/Domestic'][row] == 'Import':
                    import_SO += int(float(current_sku['Quantity Remaining'][row]))*-1
                else:
                    domestic_SO += int(float(current_sku['Quantity Remaining'][row]))*-1
            
            elif current_sku['Type Column'][row] == 'P.O.' or current_sku['Type Column'][row] == 'Non-Confirmed Blanket P.O.':
                import_PO += int(float(current_sku['Quantity Remaining'][row]))
            
            elif current_sku['Type Column'][row] == 'Confirmed P.O.' or current_sku['Type Column'][row] == 'Non-Confirmed P.O.':
                domestic_PO += int(float(current_sku['Quantity Remaining'][row]))

            elif current_sku['Type Column'][row] == 'Est. Transit Receive':
                transit += int(float(current_sku['Quantity Remaining'][row]))
            
            elif current_sku['Import/Domestic'][row] == 'Import': 
                import_forecast += int(float(current_sku['Forecast'][row]))*-1
            
            else:
                domestic_forecast += int(float(current_sku['Forecast'][row]))*-1

            """
            try:
                location_in_transit += int(float(current_sku['Location in Transit'][row]))
            except ValueError:
                pass 
            """

        try:
            available = onhand - domestic_SO
            if available < 0: 
                available = 0
        except:
            pass

        #Asign Summary Numbers to Cells
        ws.cell(row=3,column=18,value='Total Onhand:').font = double_underline
        ws.cell(row=4,column=18,value='Domestic Sales on Order:').font = double_underline
        ws.cell(row=5,column=18,value='Import Sales on Order:').font = double_underline
        ws.cell(row=6,column=18,value='Domestic Purchase on Order:').font = double_underline
        ws.cell(row=7,column=18,value='Import Purchase on Order:').font = double_underline
        ws.cell(row=8,column=18,value='Transit:').font = double_underline
        ws.cell(row=9,column=18,value='Available:').font = double_underline
        ws.cell(row=10,column=18,value='Domestic Forecast:').font = double_underline
        ws.cell(row=11,column=18,value='Import Forecast:').font = double_underline

        ws.cell(row=3,column=18).alignment = right
        ws.cell(row=4,column=18).alignment = right
        ws.cell(row=5,column=18).alignment = right
        ws.cell(row=6,column=18).alignment = right
        ws.cell(row=7,column=18).alignment = right
        ws.cell(row=8,column=18).alignment = right
        ws.cell(row=9,column=18).alignment = right
        ws.cell(row=10,column=18).alignment = right
        ws.cell(row=11,column=18).alignment = right

        ws.cell(row=3,column=19,value=onhand).alignment = center
        ws.cell(row=4,column=19,value=domestic_SO).alignment = center
        ws.cell(row=5,column=19,value=import_SO).alignment = center
        ws.cell(row=6,column=19,value=domestic_PO).alignment = center
        ws.cell(row=7,column=19,value=import_PO).alignment = center
        ws.cell(row=8,column=19,value=transit).alignment = center
        ws.cell(row=9,column=19,value=available).alignment = center
        ws.cell(row=10,column=19,value=domestic_forecast).alignment = center
        ws.cell(row=11,column=19,value=import_forecast).alignment = center


        #Stock below 0 date Calculations
        #actual_below0 = ''
        #actual_days = ''
        #forecast_below0 = ''
        #forecast_days = ''
        #for row in Wizard:
        #    if SKU == row['Item']: 
        #        try: 
        #            if float(row['Actual Inventory Level']) < 0: 
        #                actual_below0 = row['Schedule Date']
        #                actual_belowedit = re.split('-',actual_below0)
        #                actual_days = int((datetime.date(int(actual_belowedit[0]),int(actual_belowedit[1]),int(actual_belowedit[2])) - datetime.date.today()).days)
        #                break
        #        except ValueError: 
        #            pass

        #for row in Wizard:
        #    if SKU == row['Item']:
        #        try:
        #            if float(row['Forecast Inventory Level']) < 0:
        #                forecast_below0 = row['Schedule Date']
        #                forecast_belowedit = re.split('-',forecast_below0)
        #                forecast_days = int((datetime.date(int(forecast_belowedit[0]),int(forecast_belowedit[1]),int(forecast_belowedit[2])) - datetime.date.today()).days)
        #                break
        #        except ValueError:
        #            pass
        
        #ws.cell(row=3,column=11,value='Run Date:')
        #ws.cell(row=4,column=11,value='Run Time:')
        #ws.cell(row=3,column=12,value=date)
        #ws.cell(row=4,column=12,value=time)

        #ws.cell(row=10,column=7,value='Actual').border = thin_border
        #ws.cell(row=11,column=7,value='Forecast').border = thin_border
        #ws.cell(row=9,column=8,value='Stock below 0 Date').border = thin_border
        #ws.cell(row=9,column=9,value='Days').border = thin_border

        #ws.cell(row=10,column=8,value=actual_below0).border = thin_border
        #ws.cell(row=11,column=8,value=forecast_below0).border = thin_border
        #ws.cell(row=10,column=9,value=actual_days).border = thin_border
        #ws.cell(row=11,column=9,value=forecast_days).border = thin_border
        #ws.cell(row=11, column=7).font = openpyxl.styles.Font(size=9)
        #ws.cell(row=10, column=7).font = openpyxl.styles.Font(size=9)

        # Demand Multiplier
        ws.cell(row=12,column=17,value=1).number_format = FORMAT_PERCENTAGE
        ws.cell(row=12,column=15,value='Demand Multiplier')
        ws.merge_cells(start_row=12,start_column=15,end_row=12,end_column=16)

        # Lunar New Year Calander
        #ws.cell(row=3,column=3,value = 'Chinese New Year Factory Schedule').font = openpyxl.styles.Font(bold=True,italic = True,underline = 'single')
        #ws.cell(row=4,column=3, value = 'Order Date before:').font = italic
        #ws.cell(row=5,column=3, value = 'Ship Date before:').font = italic
        #ws.cell(row=6,column=3,value = 'Factory Open Date:').font = italic
        #ws.cell(row=7,column=3,value = 'Order Date After:').font = italic
        #ws.cell(row=8,column=3,value = 'Ship Date After:').font = italic
#
        #for row in Lunar:
            #if row['Factory Name'] == factory:
                #ws.cell(row=4,column=5,value  = row['Order Date before CNY'])
                #ws.cell(row=5,column=5,value = row['Ship Date before CNY'])
                #ws.cell(row=6,column=5,value = row['Factory Open Date'])
                #ws.cell(row=7,column=5,value = row['Order Date after CNY'])
                #ws.cell(row=8,column=5,value = row['Ship Date after CNY'])
        
    
    #save workbook
    wb.save(book_name)

    # SMTP server configuration
    smtp_server = "smtp.local.blackstoneproducts.com"  
    smtp_port = 25  

    # Content
    sender_email = "keaton@blackstoneproducts.com"
    subject = 'Inventory Wizard - {:%m/%d/%y}'.format(datetime.date.today())
    file_path = r'C:\Users\Keaton\Documents\IW4\\' + book_name
    file_name = book_name

    # Create the email
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = ", ".join(email_list)
    msg["Subject"] = subject
    body = """Hi all,

    See attached for today's Inventory Wizard.
    

    Best,

    Keaton Manwaring
    """

    # Attach the email body
    msg.attach(MIMEText(body, "plain"))
    
    try:
        with open(file_path, "rb") as file:
            # Set up the MIMEBase object for the attachment
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(file.read())
            
        # Encode the file in base64
        encoders.encode_base64(part)
        # Add the appropriate headers for the attachment
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={file_name}",
        )
        # Attach the file to the email
        msg.attach(part)
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        exit()

    # Send the email
    try:
        # Connect to the SMTP server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            # No login required since the server authenticates based on your IP
            server.sendmail(sender_email, email_list, msg.as_string())  # Send the email
            print("Email sent successfully!")
    except Exception as e:
        print(f"Error: {e}")







        
    
if __name__ == '__main__':
    main()



